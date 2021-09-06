import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from matplotlib import colors
from openpyxl.utils import get_column_letter
import os
from colorama import Fore, Style


def save(bP, sourceFilename, colorscaleOutput):
    """ Builds and saves (as an Excel file) a visual for the physical assembly of an output image.

    1. For each individual bead: The color name, brand, and product code are stored in the appropriate location in the Excel file. 
    2. The saved Excel file is opened and styling is applied. This is performed using openpyxl and overwrites the saved Excel file.
    """

    ### Convert beadList to a 1D array
    array1D = np.array(bP.beadList)
    
    for bead in range(len(array1D)):
        array1D[bead] = array1D[bead].color
    
    # Convert the 1D array to a 2D array
    array2D = array1D.reshape(bP.beadRows, bP.beadColumns)

    mapDF = pd.DataFrame(array2D)
    
    ### Save the bead map in Excel. If a folder associated with the image title doesn't yet exist, a new folder is created.
    outputTitle = sourceFilename[sourceFilename.find("/") + 1:sourceFilename.find(".")] # Get the title of the source image

    if colorscaleOutput:
        filename = outputTitle + "_" + str(bP.totalBeads) + "_beads_MAP_COLOR" + ".xlsx"
    else:
        filename = outputTitle + "_" + str(bP.totalBeads) + "_beads_MAP_GRAY" + ".xlsx"

    if os.path.isdir("Outputs" + "/" + outputTitle + "/"):
        mapDF.to_excel("Outputs" + "/" + outputTitle + "/" + filename)
    else:
        os.makedirs("Outputs" + "/" + outputTitle + "/")
        mapDF.to_excel("Outputs" + "/" + outputTitle + "/" + filename)
  

    ### Change cell colors
    wb = load_workbook("Outputs" + "/" + outputTitle + "/" + filename)
    wb = applyStyling(wb, bP, array2D)
    
    
    wb.save("Outputs" + "/" + outputTitle + "/" + filename)
    
    
    print(Fore.GREEN + "--------------------------------------------------")
    print("Bead map SAVED" + Style.RESET_ALL)

    return "Outputs" + "/" + outputTitle + "/" + filename
    

def applyStyling(wb, bP, array2D):
    """ Applies styling to the bead map Excel file

    Styling includes cell coloring and changes to column width.
    """

    sheet = wb.active

    ### Build dictionary with key, value = RGB, (color name, brand, product code).
    rgbToBrandAndProductCode = {}
    for bead in bP.beadList:
        rgbToBrandAndProductCode[bead.color] = bead.matchedColorName + " | " + bead.matchedColorBrand + " | " + bead.matchedColorProductCode


    ### Using the dictionary, loop through all cells and change their fill color to the corresponding RGB color.
    shape = array2D.shape
    width = shape[0]
    height = shape[1]

    for row in range(width):
        for col in range(height):

            r = array2D[row][col][0]
            g = array2D[row][col][1]
            b = array2D[row][col][2]

            sheet[row + 2][col+1].value = rgbToBrandAndProductCode[(r,g,b)]

            r = array2D[row][col][0] / 255
            g = array2D[row][col][1] / 255
            b = array2D[row][col][2] / 255

            hex = colors.rgb2hex((r, g, b))

            hex = hex[1:]

            sheet[row+2][col+1].fill = PatternFill(fgColor=hex, fill_type = "solid")
    
    
    ### Modify the Excel column widths so that cells take on small, square shape.
    ### For reference, column_dimensions[i].width = 2.7 is the same as a column width of 180 pixels.
    column = 1
    while column < bP.beadColumns + 2:
        i = get_column_letter(column)
        sheet.column_dimensions[i].width = 2.7
        column += 1

    
    return wb