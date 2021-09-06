import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from matplotlib import colors
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
from collections import Counter
from colorama import Fore, Style


def save(bP, sourceFilename, colorscaleOutput, pegboardDimensions, pegboardsRequired):
    """ Builds and saves (as an Excel file) a table listing the frequency of each bead.
    
    1. For each unique bead: The RGB values, name, brand, and product code are stored. This is performed with Pandas and saved in Excel.
    2. The saved Excel file is opened and styling is applied. This is performed using openpyxl and overwrites the saved Excel file.
    """

    ### Get the title of the source image.
    outputTitle = sourceFilename[sourceFilename.find("/") + 1:sourceFilename.find(".")]
    
    ### Builds a Counter object which reports the frequency of each unique bead color.
    counter = getCounter(bP)

    ### Populate DataFrame with bead information. Only information is added with this function; no styling is applied.
    returned = populateDF(counter, bP.totalBeads, pegboardDimensions, pegboardsRequired)
    bead_DF = returned[0]
    rgbs = returned[1]

    ### Set the save filename. Filename is specified for colorscale or grayscale output.
    if colorscaleOutput:
        filename = outputTitle + "_" + str(bP.totalBeads) + "_beads_COUNTER_COLOR" + ".xlsx"
    else:
        filename = outputTitle + "_" + str(bP.totalBeads) + "_beads_COUNTER_GRAY" + ".xlsx"    
    
    ### Save the bead counter in Excel. If a folder associated with the image title doesn't yet exist, a new folder is created.
    outputTitle = sourceFilename[sourceFilename.find("/") + 1:sourceFilename.find(".")] # Get the title of the source image
    
    if os.path.isdir("Outputs" + "/" + outputTitle + "/"):
        bead_DF.to_excel("Outputs" + "/" + outputTitle + "/" + filename)
    else:
        os.makedirs("Outputs" + "/" + outputTitle + "/")
        bead_DF.to_excel("Outputs" + "/" + outputTitle + "/" + filename)
    

    ### Now open the saved Excel file and apply styling to the workbook via openpyxl.
    wb = load_workbook("Outputs" + "/" + outputTitle + "/" + filename)
    wb = applyStyling(wb, rgbs)
    wb.save("Outputs" + "/" + outputTitle + "/" + filename)

    print(Fore.GREEN + "--------------------------------------------------")
    print("Bead counter SAVED" + Style.RESET_ALL)

    return "Outputs" + "/" + outputTitle + "/" + filename


def getCounter(bP):
    """ Builds a Counter object which reports the frequency of each unique bead color.
    """

    beadInfo = []

    ### The unique info for each bead is composed of color (RGB), color name, brand, and product code.
    for bead in bP.beadList:
        info = (bead.color, bead.matchedColorName, bead.matchedColorBrand, bead.matchedColorProductCode)
        beadInfo.append(info)

    return Counter(beadInfo)


def populateDF(counter, totalBeads, pegboardDimensions, pegboardsRequired):

    keys = list(counter.keys())

    rgbs = []
    names = []
    brands = []
    codes = []

    ### For each key (each unique bead), append its info to a specific list.
    for k in keys:
        rgbs.append(k[0])
        names.append(k[1])
        brands.append(k[2])
        codes.append(k[3])

    ### Build and populate the DataFrame with bead-specific information.
    df = pd.DataFrame()
    df["RGB"] = rgbs
    df["Color"] = [""] * len(counter) # Intentionally left blank. Will be color-filled later.
    df["Name"] = names
    df["Brand"] = brands
    df["Product Code"] = codes
    df["Count"] = counter.values()

    ### Whitespace columns serve to visually divide the left and right sides of the Excel sheet.
    df["Whitespace Col1"] = [""] * len(counter)
    df["Whitespace Col2"] = [""] * len(counter)

    ### Insert a single-row column specifying the total number of beads in the output image.
    temp1 = [""] * len(counter)
    temp1[0] = totalBeads
    df["Total Beads"] = temp1

    ### Insert a single-row column specifying the dimensions of the selected pegboard.
    temp2 = [""] * len(counter)
    temp2[0] = str(pegboardDimensions[0]) + ", " + str(pegboardDimensions[1])
    df["Pegboard Dimensions"] = temp2

    ### Insert a single-row column specifying the number of pegboards required to build the output image.
    temp3 = [""] * len(counter)
    temp3[0] = pegboardsRequired
    df["Pegboards Required"] = temp3

    return (df, rgbs)


def applyStyling(wb, rgbs):
    """ Applies styling to the counter Excel file

    Styling includes cell coloring, changes to column width, and border adjustments.
    """

    ### Set as active sheet from the workbook
    sheet = wb.active

    ### Modify the Excel column widths so that columns don't look crammed upon opening the file.
    ### For reference, column_dimensions[i].width = 20 is the same as a column width of 180 pixels.
    i = get_column_letter(2)
    sheet.column_dimensions[i].width = 14 # "RGB" column
    i = get_column_letter(4)
    sheet.column_dimensions[i].width = 18 # "Name" column
    i = get_column_letter(6)
    sheet.column_dimensions[i].width = 13 # "Product Code" column
    i = get_column_letter(10)
    sheet.column_dimensions[i].width = 15 # "Total Beads" column
    i = get_column_letter(11)
    sheet.column_dimensions[i].width = 20 # "Pegboards Dimensions" column
    i = get_column_letter(12)
    sheet.column_dimensions[i].width = 20 # "Pegboards Required" column

    ### For each row, change the Color column's fill color to match the row's RGB value.
    for row in range(sheet.max_row - 1):
        r = rgbs[row][0] / 255
        g = rgbs[row][1] / 255
        b = rgbs[row][2] / 255
        
        hex = colors.rgb2hex((r, g, b))
        hex = hex[1:]
        sheet[row + 2][2].fill = PatternFill(fgColor=hex, fill_type = "solid")


    ### To visually separate the left and right side of the Excel sheet, two columns' borders are removed and their labels are reset to empty strings.
    style = Side(border_style=None)
    border = Border(top=style, bottom=style,right=style,left=style)
    sheet[1][7].border = border
    sheet[1][7].value = ""
    sheet[1][8].border = border
    sheet[1][8].value = ""


    return wb
